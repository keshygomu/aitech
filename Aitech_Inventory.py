import streamlit as st
import pandas as pd
#from consumer_details import DOMAIN, USERNAME, PASSWORD, CONSUMER_KEY, CONSUMER_SECRET
import os
from datetime import datetime
import requests
from openpyxl import load_workbook, Workbook
from simple_salesforce import Salesforce


st.image('aitech_logo.png', use_column_width=True)
# Título da aplicação
#st.subheader("棚卸入力フォーム")

# Inicializa as variáveis no session_state para controlar os valores dos inputs
if "botao_confirmar_ativo" not in st.session_state:
    st.session_state.botao_confirmar_ativo = True


# Gera uma chave única para cada campo
def get_key(base):
    return f"{base}_{st.session_state.botao_confirmar_ativo}"


# Função para autenticar e criar uma instância do Salesforce
def authenticate_salesforce():
    auth_url = f"{st.secrets[DOMAIN]}/services/oauth2/token"
    auth_data = {
        'grant_type': 'password',
        'client_id': st.secrets[CONSUMER_KEY],
        'client_secret': st.secrets[CONSUMER_SECRET],
        'username': st.secrets[USERNAME],
        'password': st.secrets[PASSWORD]
    }
    response = requests.post(auth_url, data=auth_data)
    response.raise_for_status()
    access_token = response.json()['access_token']
    instance_url = response.json()['instance_url']
    return Salesforce(instance_url=instance_url, session_id=access_token)


# Mapeamento dos status
status_mapping = {
    "BeforeOrderConfirmation": "確定前",
    "OrderConfirmed": "確定",
    "InProduction": "製造中",
    "Done": "作業完了",
    "Cancelled": "キャンセル"
}

# Campo de entrada para o código (texto)
codigo_input_id = get_key("codigo_input")
codigo_input = st.text_input(
    "移行票番号を入力してください:",  # Label alterado
    key=codigo_input_id
)

# Formata o código no formato "PO-000000"
codigo_formatado = f"PO-{int(codigo_input):06d}" if codigo_input.isdigit() else None

codigo_existente = False
df_existente=[]
primeira_contagem = []
toda_contagem=[]

# Verifica se o arquivo Excel existe e faz a checagem
if codigo_formatado:
    nome_arquivo = f"棚卸_{datetime.now().strftime('%Y%m%d')}.xlsx"
    total_prodorder = 0
    total_prodorder_check = 0

    if os.path.exists(nome_arquivo):
        df_existente = pd.read_excel(nome_arquivo, sheet_name=0)
        total_prodorder = len(df_existente)
        total_prodorder_check = df_existente['時間2'].count()
        # Lista dos registros que ainda nao foram checados
        primeira_contagem = df_existente[df_existente['時間2'].isna()]
        # Verifica se o código já existe na coluna B
        codigo_existente = codigo_formatado in df_existente['移行票№'].values
        if codigo_existente:
            st.warning("登録済み")  # Exibe a mensagem de alerta

# Realiza a consulta ao Salesforce ao inserir o código
if codigo_input:
    try:
        sf = authenticate_salesforce()
        query = f"""
        SELECT Name, snps_um__ProcessName__c, snps_um__ActualQt__c, snps_um__Item__r.Name, 
               snps_um__Item__r.AITC_PrintItemName__c, snps_um__ProcessOrderNo__c, 
               snps_um__ProdOrder__r.Name, snps_um__Status__c, snps_um__WorkPlace__r.Name,
               snps_um__StockPlace__r.Name, snps_um__Item__c, snps_um__Process__r.Process_cost__c, 
               snps_um__Item__r.AITC_ItemRank__c, snps_um__Item__r.snps_um__Weight__c, 
               AITC_OrderQt__c 
        FROM snps_um__WorkOrder__c 
        WHERE snps_um__ProdOrder__r.Name = '{codigo_formatado}'
        """
        result = sf.query(query)

        if result['totalSize'] > 0:
            # Exibe os itens que não se repetem no topo (considerando o primeiro registro)
            first_record = result['records'][0]
            prod_order_no = first_record['snps_um__ProdOrder__r']['Name']
            item_name = first_record['snps_um__Item__r']['Name']
            item_print_name = first_record['snps_um__Item__r']['AITC_PrintItemName__c']
            rank = first_record['snps_um__Item__r']['AITC_ItemRank__c']
            weight = first_record['snps_um__Item__r']['snps_um__Weight__c']
            original_order = first_record['AITC_OrderQt__c']

            st.write(f"**移行票№**: {prod_order_no}　ー　{original_order}")
            st.write(f"**品目**: {item_name}　**品名**: {item_print_name}　**ランク**: {rank}　**重量**:　{weight}")

            # Cria a tabela para exibir os dados
            table_data = []
            headers = ["作業オーダー", "工程", "順序", "数量", "ステータス", "作業場所", "工程単価"]
            for record in result['records']:
                process_name = record['snps_um__ProcessName__c']
                process_order_no = int(record['snps_um__ProcessOrderNo__c'])
                status = status_mapping.get(record['snps_um__Status__c'], record['snps_um__Status__c'])
                actual_qty = int(record['snps_um__ActualQt__c'])
                work_place_name = record['snps_um__WorkPlace__r']['Name']  # Extraído de cada registro
                cost_price = record['snps_um__Process__r']['Process_cost__c']

                table_data.append([
                    record['Name'],  # 作業オーダー
                    process_name,  # 工程
                    process_order_no,  # 順序, sem casas decimais
                    actual_qty,  # 数量, sem casas decimais
                    status,  # ステータス traduzido
                    work_place_name,   # 作業場所
                    str(round(cost_price,2)) # 工程単価
                ])

            # Cria o DataFrame
            df = pd.DataFrame(table_data, columns=headers)

            # Filtra o último valor maior que 0 da coluna "数量"
            try:
                last_non_zero_quantity = df[df['数量'] > 0].iloc[-1]  # Filtra e seleciona a última linha
                last_line = int(last_non_zero_quantity.name)
                acum_price = 0
                x = 0
                for record in result['records']:
                    if x <= last_line:
                        acum_price = acum_price + float(record['snps_um__Process__r']['Process_cost__c'])
                        x = x + 1
            except:
                last_non_zero_quantity = None
                acum_price = 0


            # Aplica formatação condicional
            def highlight_zero_quantity(row):
                return ['background-color: lightgreen' if row['数量'] != 0 else '' for _ in row]


            # Aplica a formatação ao DataFrame e exibe a tabela no Streamlit
            styled_df = df.style.apply(highlight_zero_quantity, axis=1)
            with st.popover("製造オーダー明細"):
                st.dataframe(styled_df)
                st.text(f"工程終了までの単価:  　{str(round(acum_price,3))}")

        else:
            st.warning("入力されたコードに対して、レコードが見つかりませんでした。")  # Aviso traduzido
            last_non_zero_quantity = None  # Caso não encontre, não retorna uma linha
    except Exception as e:
        st.error(f"Salesforceへの問い合わせでエラーが発生しました: {str(e)}")  # Erro traduzido
        last_non_zero_quantity = None
else:
    last_non_zero_quantity = None


# Campo de entrada para a quantidade (texto), preenchido com o último valor maior que 0
quantidade_input_id = get_key("quantidade_input")
quantidade = st.text_input(
    "数量:", max_chars=10,
    value=str(last_non_zero_quantity['数量']) if last_non_zero_quantity is not None else "0",
    key=quantidade_input_id
)

# Campo de entrada para o código do responsável (texto)
codigo_responsavel_input_id = get_key("codigo_responsavel_input")
codigo_responsavel = st.text_input(
    "担当者コード",  # Label alterado
    key=codigo_responsavel_input_id
)

# Verificação se todos os campos estão preenchidos
botao_confirmar_ativado = st.session_state.botao_confirmar_ativo and codigo_input and quantidade and codigo_responsavel

# Função para salvar os dados em um arquivo Excel
def salvar_dados_excel(codigo, quantidade, codigo_responsavel, last_non_zero_quantity, cost_price):
    # Formata o nome do arquivo Excel com a data atual
    data_atual = datetime.now().strftime("%Y%m%d")
    nome_arquivo = f"棚卸_{data_atual}.xlsx"

    if os.path.exists(nome_arquivo):
        # Carrega o arquivo existente
        workbook = load_workbook(nome_arquivo)
        sheet = workbook.active
        df_existente = pd.read_excel(nome_arquivo)

        # Verifica se o código já existe na coluna B
        if codigo_formatado in df_existente['移行票№'].values:
            row_index = df_existente.index[df_existente['移行票№'] == codigo_formatado][0]+2

            # Encontra a próxima coluna vazia na linha correspondente
            col_index = df_existente.columns.get_loc('移行票№') + 1  # Começa após a coluna '移行票№'
            while sheet.cell(row=row_index, column=col_index).value is not None:
                col_index += 1
            # Escreve os novos dados na próxima coluna vazia
            sheet.cell(row=row_index, column=col_index, value=datetime.now().strftime("%H:%M:%S"))  # Horário
            sheet.cell(row=row_index, column=col_index + 1, value=quantidade)  # Quantidade
            sheet.cell(row=row_index, column=col_index + 2, value=codigo_responsavel)  # Código do Responsável

        else:
            # Adiciona nova linha se o código não existir
            sheet.append([datetime.now().strftime("%H:%M:%S"),
                          codigo_formatado,
                          int(quantidade),
                          int(codigo_responsavel),
                          item_name, last_non_zero_quantity.get('工程', ''),
                          int(last_non_zero_quantity.get('順序', '')),
                          last_non_zero_quantity.get('作業場所', ''),
                          cost_price])
    else:
        # Cria um novo arquivo e adiciona colunas extra
        workbook = Workbook()
        sheet = workbook.active
        # Define os cabeçalhos das colunas
        colunas = ['時間', '移行票№', '数量', '担当者', '品目', '工程', '順序', '作業場所',
                   '累積コスト', '時間2', '数量2', '担当者2', '時間3', '数量3', '担当者3']
        sheet.append(colunas)

        # Adiciona os dados na nova linha
        sheet.append([datetime.now().strftime("%H:%M:%S"),
            codigo_formatado, int(quantidade),
            int(codigo_responsavel),
            item_name,
            last_non_zero_quantity.get('工程', ''),
            int(last_non_zero_quantity.get('順序', '')),
            last_non_zero_quantity.get('作業場所', ''),
            cost_price])
    workbook.save(nome_arquivo)


# Botão de confirmação da entrada de dados
if st.button("データ登録", disabled=not botao_confirmar_ativado, type="primary"):  # Texto do botão alterado
    try:
        # Salva os dados no arquivo Excel
        salvar_dados_excel(codigo_input, quantidade, codigo_responsavel, last_non_zero_quantity, acum_price)
        if codigo_existente:
            if total_prodorder > total_prodorder_check:
                st.warning(f"移行票 {total_prodorder}件(登録済み)　再確認 {total_prodorder_check + 1}件")
            else:
                st.warning(f"移行票 {total_prodorder}件(登録済み)　再確認 {total_prodorder_check}件")
        else:
            st.warning(f"移行票 {total_prodorder+1}件(登録済み)　再確認 {total_prodorder_check}件")
        st.success("データが正常に確認されました！")  # Mensagem de sucesso traduzida
        st.write(f"移行票№: {codigo_formatado} / {item_name}")  # Código formatado e label atualizado
        st.write(f"数量: {quantidade}     担当者コード: {codigo_responsavel}")  # Label atualizado
    except:
        st.write(f"生産が開始されていないため。移行票№: {codigo_formatado}　は登録されません。") # Nao ha valores maiores que 0

    # Desabilita o botão de confirmação até uma nova entrada ser feita
    st.session_state.botao_confirmar_ativo = False

col1, col2 = st.columns(2)

with col1:
    with st.popover("再確認待ち"):
        st.dataframe(primeira_contagem)
with col2:
    with st.popover("現在棚卸詳細"):
        st.dataframe(df_existente)

# Reativa o botão de confirmação quando o usuário começar a digitar em qualquer campo
if not st.session_state.botao_confirmar_ativo:
    st.session_state.botao_confirmar_ativo = True
